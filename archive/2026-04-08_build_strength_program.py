"""
Program Builder V2
==================
Flexible multi-progression Excel program generator + patcher.

Usage:
    python src/build_strength_program.py                           # Nacho's 30-week program
    python src/build_strength_program.py --demo                    # 4 sample files, one per theme
    python src/build_strength_program.py --extend-current FILE     # add RPE + Date to existing file
    python src/build_strength_program.py --theme dark_minimal      # change theme for default program
    python src/build_strength_program.py --output DIR              # output directory (default: .)

Integration: import BUILDER_TOOL_MANIFEST and MODIFIER_TOOL_MANIFEST in program_agent.py.
Requirements: pip install openpyxl
"""
from __future__ import annotations

import argparse
import os
import re
from dataclasses import dataclass, field
from datetime import date, timedelta
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# TOOL MANIFESTS  (imported by program_agent.py to inject into coach prompt)
# ─────────────────────────────────────────────────────────────────────────────

BUILDER_TOOL_MANIFEST = {
    "tool": "ProgramBuilder",
    "description": (
        "Generates a complete Excel strength program workbook from a ProgramConfig. "
        "Creates Overview, Progression Plan, all Week tabs, and a Rules sheet. "
        "Returns the path to the generated .xlsx file."
    ),
    "when_to_use": (
        "Use when CREATE_NEW is decided and a structured Excel file is wanted. "
        "The file must then be uploaded to replace the Google Sheet."
    ),
    "required_inputs": {
        "name": "str — program name",
        "total_weeks": "int — program length",
        "days": "list[DayConfig] — day definitions with exercises",
        "progression_method": (
            "str — one of: LINEAR, STEP_LOAD, DOUBLE_PROGRESSION, WAVE, "
            "DUP_DAILY, BLOCK_PERIODIZATION, PERCENTAGE_BASED, RPE_BASED"
        ),
    },
    "optional_inputs": {
        "start_date": "date — auto-fills session dates in day headers",
        "blocks": "list[BlockConfig] — rep scheme evolution per block",
        "pullback_weeks": "list[int] — explicit deload weeks",
        "theme": "str — one of: professional_blue, dark_minimal, warm_earth, clean_forest",
        "include_cardio": "bool — adds Cardio section (default False)",
        "include_wildcard_slots": "bool — blank surprise session rows every N weeks (default False)",
        "wildcard_frequency": "int — wildcard every N weeks (default 4)",
        "rpe_column": "bool — add RPE column (default True)",
        "notes_column": "bool — add Session Notes column (default True)",
        "pullback_pct": "float — deload weight % (default 0.90)",
        "rounding": "float — round weights to nearest X kg (default 2.5)",
    },
    "available_progressions": [
        "LINEAR — fixed kg increment per effective week",
        "STEP_LOAD — 3 weeks load + 1 week deload, auto-computed pullbacks",
        "DOUBLE_PROGRESSION — hold weight until rep range top hit, then add",
        "WAVE — 3-week repeating heavy/moderate/volume cycle",
        "DUP_DAILY — different rep scheme each day (strength/hypertrophy/power/volume)",
        "BLOCK_PERIODIZATION — accumulation -> intensification -> realization",
        "PERCENTAGE_BASED — % of training max (1RM); shows computed kg + %",
        "RPE_BASED — no prescribed weights; shows RPE targets only",
    ],
    "available_themes": ["professional_blue", "dark_minimal", "warm_earth", "clean_forest"],
    "example_call": (
        "from src.build_strength_program import ProgramBuilder, ProgramConfig, DayConfig, ExerciseConfig\n"
        "cfg = ProgramConfig(name='12-Week Hypertrophy', total_weeks=12,\n"
        "    days=[DayConfig(label='Day 1', focus='Upper Push',\n"
        "        exercises=[ExerciseConfig(name='Bench Press', lift_key='bench', "
        "start_weight=75, increment=1.25)])],\n"
        "    progression_method='DOUBLE_PROGRESSION', theme='clean_forest')\n"
        "path = ProgramBuilder(cfg).build('output/new_program.xlsx')"
    ),
}

MODIFIER_TOOL_MANIFEST = {
    "tool": "ProgramModifier",
    "description": (
        "Patches an existing Excel program file without recreating it. "
        "All methods return self for chaining. Call .save() at the end."
    ),
    "when_to_use": (
        "Use for MODIFY_CURRENT operations that affect a week range or need structural "
        "changes: scaling weights, rep scheme changes, swaps, adding columns."
    ),
    "operations": {
        "scale_weights(week_range, pct)": "Scale all numeric weights to pct% (e.g., 90.0 = 90%)",
        "change_rep_scheme(week_range, exercise, new_scheme)": "Change scheme string for one exercise",
        "swap_exercise(old, new, week_range)": "Rename an exercise across tabs",
        "add_rpe_column(week_range)": "Insert RPE column after Actual if not present",
        "add_date_column(week_range, start_date, day_offsets)": (
            "Append date text to day headers; auto-fills if start_date given"
        ),
        "add_wildcard_slot(after_week, num_rows)": "Add blank wildcard rows at bottom of a week tab",
        "retheme(theme, week_range)": "Apply new color theme to all header rows",
        "save(path)": "Save; defaults to overwriting original file",
    },
    "example_call": (
        "from src.build_strength_program import ProgramModifier\n"
        "from datetime import date\n"
        "ProgramModifier('strength_30weeks.xlsx')\\\n"
        "    .add_rpe_column(week_range=(10, 30))\\\n"
        "    .add_date_column(week_range=(10, 30), start_date=date(2026, 1, 13))\\\n"
        "    .save('strength_30weeks_v2.xlsx')"
    ),
}

# ─────────────────────────────────────────────────────────────────────────────
# THEME SYSTEM
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class Theme:
    name: str
    header_bg: str          # column-label header background
    header_fg: str          # column-label header font color
    day_bg: str             # day section header background
    day_fg: str             # day section header font color
    title_font: str         # font name for titles
    title_size: int
    body_font: str          # font name for data rows
    body_size: int
    pullback_bg: str        # deload week highlight
    test_bg: str            # test / peak week highlight
    wildcard_bg: str        # wildcard slot highlight
    alt_row_bg: str         # alternating row tint ("" = none)
    border_style: str       # "thin" | "medium" | "none"


THEMES: dict[str, Theme] = {
    "professional_blue": Theme(
        name="Professional Blue",
        header_bg="2F5496", header_fg="FFFFFF",
        day_bg="4472C4", day_fg="FFFFFF",
        title_font="Calibri", title_size=14,
        body_font="Calibri", body_size=11,
        pullback_bg="90EE90", test_bg="FFD700", wildcard_bg="E8E8FF",
        alt_row_bg="EEF4FF", border_style="thin",
    ),
    "dark_minimal": Theme(
        name="Dark Minimal",
        header_bg="1F2937", header_fg="F9FAFB",
        day_bg="374151", day_fg="F9FAFB",
        title_font="Consolas", title_size=13,
        body_font="Consolas", body_size=10,
        pullback_bg="1E3A2F", test_bg="3B2F00", wildcard_bg="1E1E2E",
        alt_row_bg="2D3748", border_style="thin",
    ),
    "warm_earth": Theme(
        name="Warm Earth",
        header_bg="6B3A2A", header_fg="FFF8F0",
        day_bg="8B5E3C", day_fg="FFF8F0",
        title_font="Georgia", title_size=14,
        body_font="Georgia", body_size=11,
        pullback_bg="D4A96A", test_bg="B8860B", wildcard_bg="F5DEB3",
        alt_row_bg="FDF5E6", border_style="thin",
    ),
    "clean_forest": Theme(
        name="Clean Forest",
        header_bg="1B5E20", header_fg="F1F8E9",
        day_bg="2E7D32", day_fg="F1F8E9",
        title_font="Calibri", title_size=14,
        body_font="Calibri", body_size=11,
        pullback_bg="A5D6A7", test_bg="FFE082", wildcard_bg="C8E6C9",
        alt_row_bg="F1F8E9", border_style="thin",
    ),
}

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG DATACLASSES
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class ExerciseConfig:
    """A single exercise within a training day."""
    name: str
    lift_key: str                           # internal identifier ("squat", "bench", …)
    start_weight: float = 0.0              # week-1 weight in kg (eff_weeks=0)
    increment: float = 0.0                 # kg added per effective week
    # Alternative: weight as % of another lift's weight
    ref_lift: Optional[str] = None         # if set, weight = ref_lift_weight × ref_pct
    ref_pct: float = 1.0
    # Scheme
    override_scheme: Optional[str] = None  # always use this regardless of block
    # Display
    weight_format: str = "kg"              # "kg" | "+kg" (added weight) | "bw" | "pct"
    form_cue: str = ""
    is_cardio: bool = False
    is_optional: bool = False


@dataclass
class DayConfig:
    """A training day (repeated across all weeks)."""
    label: str                              # "Day 1", "Day 2", …
    focus: str = ""                         # "Heavy Squat + Bench"
    est_time: str = ""                      # "~50 min"
    exercises: list[ExerciseConfig] = field(default_factory=list)


@dataclass
class BlockConfig:
    """A training phase spanning a range of weeks."""
    start_week: int
    end_week: int
    label: str                              # "Accumulation", "Intensification", …
    schemes: dict[str, str] = field(default_factory=dict)   # lift_key → "3x5"
    is_deload: bool = False


@dataclass
class ProgramConfig:
    """Complete program specification."""
    name: str
    total_weeks: int
    days: list[DayConfig]
    # Optional
    start_date: Optional[date] = None
    progression_method: str = "LINEAR"
    blocks: list[BlockConfig] = field(default_factory=list)
    pullback_weeks: list[int] = field(default_factory=list)
    pullback_schemes: Optional[dict[str, str]] = None   # override schemes on deload weeks
    theme: str = "professional_blue"
    include_cardio: bool = False
    include_wildcard_slots: bool = False
    wildcard_frequency: int = 4
    rpe_column: bool = True
    notes_column: bool = True
    pullback_pct: float = 0.90
    rounding: float = 2.5
    philosophy: str = ""
    goals: list[dict] = field(default_factory=list)     # [{"lift","start","goal","gain"}]
    rules_text: str = ""
    output_path: str = "."
    # Day-offset pattern for date computation (index = day position, value = days from Mon)
    day_date_offsets: list[int] = field(default_factory=lambda: [0, 1, 3, 4])

# ─────────────────────────────────────────────────────────────────────────────
# HELPER UTILITIES
# ─────────────────────────────────────────────────────────────────────────────

def _round_weight(w: float, rounding: float) -> float:
    return round(w / rounding) * rounding


def _make_fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _make_border(style: str) -> Border:
    if style == "none":
        return Border()
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)


def _parse_week_num(name: str) -> Optional[int]:
    m = re.match(r"(?i)(?:week|semana|w)\s*(\d+)$", name.strip())
    return int(m.group(1)) if m else None


def _find_header_row(ws) -> tuple[Optional[int], dict[str, int]]:
    """Find the exercise column header row. Returns (row_num, {keyword: col_index})."""
    for row_cells in ws.iter_rows():
        vals = [str(c.value or "").lower() for c in row_cells]
        joined = " ".join(vals)
        if "exercise" in joined and ("weight" in joined or "sets" in joined or "rep" in joined):
            mapping: dict[str, int] = {}
            for cell in row_cells:
                v = str(cell.value or "").lower()
                if "exercise" in v:
                    mapping["exercise"] = cell.column
                elif "weight" in v or "load" in v:
                    mapping["weight"] = cell.column
                elif "sets" in v or ("x rep" in v) or (v == "reps"):
                    mapping["sets_reps"] = cell.column
                elif v == "done":
                    mapping["done"] = cell.column
                elif "actual" in v:
                    mapping["actual"] = cell.column
                elif "rpe" in v or "effort" in v:
                    mapping["rpe"] = cell.column
                elif "session" in v or "athlete" in v:
                    mapping["session_notes"] = cell.column
                elif "note" in v or "cue" in v:
                    mapping["notes"] = cell.column
            if mapping:
                return row_cells[0].row, mapping
    return None, {}

# ─────────────────────────────────────────────────────────────────────────────
# PROGRESSION LOGIC
# ─────────────────────────────────────────────────────────────────────────────

_DUP_DAY_SCHEMES = ["4x5", "4x10", "4x3", "3x8"]
_DUP_DAY_PCT    = [1.00,   0.75,  1.05,  0.80]   # weight multiplier per day type


def _eff_weeks(week: int, pullback_weeks: list[int]) -> int:
    """Effective progression weeks (excluding deload weeks before this week)."""
    return week - 1 - sum(1 for pw in pullback_weeks if pw < week)


def _step_is_deload(week: int) -> bool:
    return (week % 4) == 0


def _step_eff_weeks(week: int) -> int:
    return week - 1 - sum(1 for w in range(1, week) if _step_is_deload(w))


def _wave_offset(eff: int) -> tuple[float, str]:
    """Return (weight_offset_kg, scheme) for WAVE method."""
    pos = eff % 3
    return [(0.0, "3x5"), (2.5, "3x3"), (-2.5, "3x6")][pos]


def _progression_weight(
    week: int,
    ex: ExerciseConfig,
    method: str,
    pullback_weeks: list[int],
    pullback_pct: float,
    rounding: float,
    day_index: int = 0,
) -> float:
    """Compute final weight for an exercise in a given week."""
    is_pb = week in pullback_weeks
    # STEP_LOAD deload uses step logic instead of pullback_weeks list
    if method == "STEP_LOAD":
        is_pb = _step_is_deload(week)
        eff = _step_eff_weeks(week)
    else:
        eff = _eff_weeks(week, pullback_weeks)

    if method == "RPE_BASED":
        return 0.0

    if method == "DOUBLE_PROGRESSION":
        # Hold weight for ~3 eff weeks before incrementing (simulates rep-range cycling)
        milestone = eff // 3
        w = ex.start_weight + milestone * ex.increment
    elif method == "WAVE":
        offset, _ = _wave_offset(eff)
        cycle_num = eff // 3
        w = ex.start_weight + cycle_num * (ex.increment * 3) + offset
    elif method == "DUP_DAILY":
        w = (ex.start_weight + eff * ex.increment) * _DUP_DAY_PCT[day_index % 4]
    else:
        w = ex.start_weight + eff * ex.increment

    if is_pb:
        w *= pullback_pct

    return _round_weight(max(w, 0), rounding)


def _progression_scheme(
    week: int,
    lift_key: str,
    method: str,
    pullback_weeks: list[int],
    pullback_schemes: Optional[dict[str, str]],
    blocks: list[BlockConfig],
    override: Optional[str] = None,
    day_index: int = 0,
) -> str:
    if override:
        return override
    if method == "STEP_LOAD" and _step_is_deload(week):
        if pullback_schemes:
            return pullback_schemes.get(lift_key, "3x5 (light)")
        return "3x5 (light)"
    if week in pullback_weeks:
        if pullback_schemes:
            return pullback_schemes.get(lift_key, "3x5")
    if method == "DUP_DAILY":
        return _DUP_DAY_SCHEMES[day_index % 4]
    if method == "WAVE":
        eff = _eff_weeks(week, pullback_weeks)
        _, scheme = _wave_offset(eff)
        return scheme
    if method == "DOUBLE_PROGRESSION":
        return "3x5-8"  # default rep range; users can override per exercise
    if method == "RPE_BASED":
        return "4x5 @RPE8"
    # BLOCK_PERIODIZATION / LINEAR / PERCENTAGE_BASED — look up from block
    for blk in blocks:
        if blk.start_week <= week <= blk.end_week:
            return blk.schemes.get(lift_key, "3x5")
    return "3x5"

# ─────────────────────────────────────────────────────────────────────────────
# PROGRAM BUILDER
# ─────────────────────────────────────────────────────────────────────────────

class ProgramBuilder:
    """Build a complete Excel program workbook from a ProgramConfig."""

    def __init__(self, config: ProgramConfig):
        self.config = config
        self.wb = Workbook()
        self.theme = THEMES.get(config.theme, THEMES["professional_blue"])
        self._border = _make_border(self.theme.border_style)
        # Pre-build weight cache (lift_key → {week: float}) for ref_lift resolution
        self._weight_cache: dict[str, dict[int, float]] = {}
        self._build_weight_cache()

    # ── Weight cache ──────────────────────────────────────────────────────────

    def _build_weight_cache(self):
        """Pre-compute weights for all non-ref exercises across all weeks."""
        for day_idx, day in enumerate(self.config.days):
            for ex in day.exercises:
                if ex.ref_lift is None:
                    self._weight_cache.setdefault(ex.lift_key, {})
                    for w in range(1, self.config.total_weeks + 1):
                        self._weight_cache[ex.lift_key][w] = _progression_weight(
                            w, ex,
                            self.config.progression_method,
                            self.config.pullback_weeks,
                            self.config.pullback_pct,
                            self.config.rounding,
                            day_index=day_idx,
                        )

    def _get_weight(self, week: int, ex: ExerciseConfig, day_index: int = 0) -> float:  # noqa: ARG002
        if ex.ref_lift:
            ref_w = self._weight_cache.get(ex.ref_lift, {}).get(week, 0.0)
            return _round_weight(ref_w * ex.ref_pct, self.config.rounding)
        return self._weight_cache.get(ex.lift_key, {}).get(week, 0.0)

    def _format_weight(self, w: float, fmt: str, method: str) -> str:
        if method == "RPE_BASED":
            return "by feel"
        if w == 0 and fmt not in ("+kg", "bw"):
            return "—"
        if fmt == "+kg":
            return f"+{w:.4g}kg"
        if fmt == "bw":
            return "BW"
        return f"{w:.4g}kg"

    def _get_scheme(self, week: int, ex: ExerciseConfig, day_index: int = 0) -> str:
        return _progression_scheme(
            week=week,
            lift_key=ex.lift_key,
            method=self.config.progression_method,
            pullback_weeks=self.config.pullback_weeks,
            pullback_schemes=self.config.pullback_schemes,
            blocks=self.config.blocks,
            override=ex.override_scheme,
            day_index=day_index,
        )

    def _active_block(self, week: int) -> Optional[BlockConfig]:
        for blk in self.config.blocks:
            if blk.start_week <= week <= blk.end_week:
                return blk
        return None

    def _week_label(self, week: int) -> str:
        is_pb = week in self.config.pullback_weeks
        is_step_dl = (self.config.progression_method == "STEP_LOAD" and _step_is_deload(week))
        is_last = (week == self.config.total_weeks)
        blk = self._active_block(week)
        blk_label = f" — {blk.label}" if blk else ""
        if is_last:
            return f"WEEK {week}{blk_label} — TEST WEEK"
        if is_pb or is_step_dl:
            return f"WEEK {week}{blk_label} — PULLBACK"
        return f"WEEK {week}{blk_label}"

    def _is_pullback(self, week: int) -> bool:
        if week in self.config.pullback_weeks:
            return True
        if self.config.progression_method == "STEP_LOAD" and _step_is_deload(week):
            return True
        return False

    def _week_fill(self, week: int) -> Optional[PatternFill]:
        if week == self.config.total_weeks:
            return _make_fill(self.theme.test_bg)
        if self._is_pullback(week):
            return _make_fill(self.theme.pullback_bg)
        return None

    def _session_date(self, week: int, day_index: int) -> Optional[date]:
        if not self.config.start_date:
            return None
        offsets = self.config.day_date_offsets
        week_start = self.config.start_date + timedelta(weeks=week - 1)
        day_offset = offsets[day_index] if day_index < len(offsets) else day_index
        return week_start + timedelta(days=day_offset)

    # ── Style helpers ──────────────────────────────────────────────────────────

    def _apply_header(self, cell, text: str, bold: bool = True):
        t = self.theme
        cell.value = text
        cell.font = Font(bold=bold, color=t.header_fg, size=t.body_size, name=t.body_font)
        cell.fill = _make_fill(t.header_bg)
        cell.border = self._border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def _apply_day_header(self, cell, text: str):
        t = self.theme
        cell.value = text
        cell.font = Font(bold=True, color=t.day_fg, size=t.body_size + 1, name=t.body_font)
        cell.fill = _make_fill(t.day_bg)

    def _apply_title(self, cell, text: str, fill: Optional[PatternFill] = None):
        t = self.theme
        cell.value = text
        cell.font = Font(bold=True, size=t.title_size, name=t.title_font)
        if fill:
            cell.fill = fill

    def _apply_data(self, cell, value, alt: bool = False, week_fill: Optional[PatternFill] = None):
        cell.value = value
        cell.border = self._border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if week_fill:
            cell.fill = week_fill
        elif alt and self.theme.alt_row_bg:
            cell.fill = _make_fill(self.theme.alt_row_bg)

    def _apply_exercise_name(self, cell, value, alt: bool = False,
                              week_fill: Optional[PatternFill] = None,
                              italic: bool = False):
        t = self.theme
        cell.value = value
        cell.border = self._border
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.font = Font(italic=italic, size=t.body_size, name=t.body_font)
        if week_fill:
            cell.fill = week_fill
        elif alt and t.alt_row_bg:
            cell.fill = _make_fill(t.alt_row_bg)

    # ── Sheet builders ─────────────────────────────────────────────────────────

    def _build_overview(self):
        ws = self.wb.active
        ws.title = "Overview"
        t = self.theme

        # Title
        ws["A1"] = self.config.name.upper()
        ws["A1"].font = Font(bold=True, size=t.title_size + 4, name=t.title_font)
        num_cols = len(self.config.days) + 1
        ws.merge_cells(f"A1:{get_column_letter(max(num_cols, 5))}1")

        # Philosophy
        if self.config.philosophy:
            ws["A3"] = "PHILOSOPHY"
            ws["A3"].font = Font(bold=True, size=12, name=t.title_font)
            for i, line in enumerate(self.config.philosophy.split("\n"), 4):
                ws.cell(row=i, column=1, value=line.strip())
            row_offset = 4 + len(self.config.philosophy.split("\n")) + 1
        else:
            row_offset = 5

        # Goals table
        if self.config.goals:
            ws.cell(row=row_offset, column=1, value="GOALS").font = Font(
                bold=True, size=12, name=t.title_font
            )
            row_offset += 1
            hdr = ["Lift", "Start", "Goal", "Total Gain"]
            for j, h in enumerate(hdr, 1):
                self._apply_header(ws.cell(row=row_offset, column=j), h)
            row_offset += 1
            for g in self.config.goals:
                ws.cell(row=row_offset, column=1, value=g.get("lift", "")).border = self._border
                ws.cell(row=row_offset, column=2, value=g.get("start", "")).border = self._border
                ws.cell(row=row_offset, column=3, value=g.get("goal", "")).border = self._border
                ws.cell(row=row_offset, column=4, value=g.get("gain", "")).border = self._border
                row_offset += 1
            row_offset += 1

        # Block structure
        if self.config.blocks:
            ws.cell(row=row_offset, column=1, value="BLOCK STRUCTURE").font = Font(
                bold=True, size=12, name=t.title_font
            )
            row_offset += 1
            for j, h in enumerate(["Block", "Weeks", "Label", "Deload?"], 1):
                self._apply_header(ws.cell(row=row_offset, column=j), h)
            row_offset += 1
            for i, blk in enumerate(self.config.blocks, 1):
                ws.cell(row=row_offset, column=1, value=i).border = self._border
                ws.cell(row=row_offset, column=2,
                        value=f"{blk.start_week}–{blk.end_week}").border = self._border
                ws.cell(row=row_offset, column=3, value=blk.label).border = self._border
                ws.cell(row=row_offset, column=4,
                        value="Yes" if blk.is_deload else "").border = self._border
                row_offset += 1
            row_offset += 1

        # Weekly day structure
        ws.cell(row=row_offset, column=1, value="WEEKLY STRUCTURE").font = Font(
            bold=True, size=12, name=t.title_font
        )
        row_offset += 1
        for j, h in enumerate(["Day", "Focus", "Est. Time", "# Exercises"], 1):
            self._apply_header(ws.cell(row=row_offset, column=j), h)
        row_offset += 1
        for day in self.config.days:
            ws.cell(row=row_offset, column=1, value=day.label).border = self._border
            ws.cell(row=row_offset, column=2, value=day.focus).border = self._border
            ws.cell(row=row_offset, column=3, value=day.est_time).border = self._border
            ws.cell(row=row_offset, column=4,
                    value=len(day.exercises)).border = self._border
            row_offset += 1
        row_offset += 1

        # Pullback / test notes
        if self.config.pullback_weeks:
            ws.cell(row=row_offset, column=1,
                    value=f"PULLBACK WEEKS: {', '.join(str(w) for w in self.config.pullback_weeks)} "
                          f"(weights ×{self.config.pullback_pct:.0%}, maintain form)").font = Font(bold=True)
            row_offset += 1
        ws.cell(row=row_offset, column=1,
                value=f"TEST WEEK: {self.config.total_weeks}").font = Font(bold=True)

        # Column widths
        for col in ["A", "B", "C", "D", "E"]:
            ws.column_dimensions[col].width = 28

    def _build_progression_sheet(self):
        ws = self.wb.create_sheet("Progression Plan")
        t = self.theme

        ws["A1"] = f"{self.config.name.upper()} — FULL PROGRESSION"
        ws["A1"].font = Font(bold=True, size=t.title_size, name=t.title_font)

        # Collect all unique lift keys from non-ref exercises
        lift_keys: list[str] = []
        lift_names: list[str] = []
        for day in self.config.days:
            for ex in day.exercises:
                if ex.ref_lift is None and ex.lift_key not in lift_keys:
                    lift_keys.append(ex.lift_key)
                    lift_names.append(ex.name)

        headers = ["Week", "Block", "Type"] + lift_names
        for j, h in enumerate(headers, 1):
            self._apply_header(ws.cell(row=3, column=j), h)

        for week in range(1, self.config.total_weeks + 1):
            row = week + 3
            blk = self._active_block(week)
            is_pb = self._is_pullback(week)
            is_test = week == self.config.total_weeks
            wfill = self._week_fill(week)

            ws.cell(row=row, column=1, value=week).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=2,
                    value=blk.label if blk else "").alignment = Alignment(horizontal="center")
            type_label = "TEST" if is_test else ("PULLBACK" if is_pb else "PROGRESS")
            ws.cell(row=row, column=3, value=type_label).alignment = Alignment(horizontal="center")

            for j, (lk, lname) in enumerate(zip(lift_keys, lift_names), 4):
                w_val = self._weight_cache.get(lk, {}).get(week, 0)
                cell = ws.cell(row=row, column=j, value=f"{w_val:.4g}kg" if w_val else "—")
                cell.alignment = Alignment(horizontal="center")
                cell.border = self._border
                if wfill:
                    cell.fill = wfill

            for col in range(1, 4):
                c = ws.cell(row=row, column=col)
                c.border = self._border
                if wfill:
                    c.fill = wfill

        for i in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 11
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 10

    def _col_headers(self) -> list[str]:
        """Exercise column headers based on config."""
        cols = ["Exercise", "Weight", "Sets x Reps", "Done", "Actual"]
        if self.config.rpe_column:
            cols.append("RPE")
        if self.config.notes_column:
            cols.append("Session Notes")
        cols.append("Notes")  # form cue — always last
        if self.config.include_cardio:
            cols.append("Cardio")
        return cols

    def _build_week_tab(self, week: int):
        ws = self.wb.create_sheet(f"Week {week}")
        t = self.theme
        wfill = self._week_fill(week)
        week_title = self._week_label(week)

        col_headers = self._col_headers()
        num_cols = len(col_headers)
        merge_end = get_column_letter(num_cols)

        # Title row
        ws["A1"] = week_title
        self._apply_title(ws["A1"], week_title, fill=wfill)
        ws.merge_cells(f"A1:{merge_end}1")
        ws["A1"].font = Font(
            bold=True, size=t.title_size,
            name=t.title_font,
            color=t.day_fg if wfill else "000000"
        )
        if wfill:
            ws["A1"].fill = wfill

        current_row = 3

        for day_idx, day in enumerate(self.config.days):
            # Day header
            d_date = self._session_date(week, day_idx)
            day_text = f"{day.label}: {day.focus}"
            if day.est_time:
                day_text += f"  ({day.est_time})"
            if d_date:
                day_text += f"  |  Date: {d_date.isoformat()}"
            self._apply_day_header(ws.cell(row=current_row, column=1), day_text)
            ws.merge_cells(f"A{current_row}:{merge_end}{current_row}")
            current_row += 1

            # Column headers
            for j, h in enumerate(col_headers, 1):
                self._apply_header(ws.cell(row=current_row, column=j), h)
            current_row += 1

            # Exercises
            for ex_idx, ex in enumerate(day.exercises):
                alt = (ex_idx % 2 == 1)
                w_val = self._get_weight(week, ex, day_idx)
                weight_str = self._format_weight(
                    w_val, ex.weight_format, self.config.progression_method
                )
                scheme = self._get_scheme(week, ex, day_idx)

                self._apply_exercise_name(
                    ws.cell(row=current_row, column=1),
                    ex.name + (" (opt)" if ex.is_optional else ""),
                    alt=alt, week_fill=wfill, italic=ex.is_optional,
                )
                self._apply_data(ws.cell(row=current_row, column=2), weight_str, alt, wfill)
                self._apply_data(ws.cell(row=current_row, column=3), scheme, alt, wfill)
                self._apply_data(ws.cell(row=current_row, column=4), "☐", alt, wfill)
                self._apply_data(ws.cell(row=current_row, column=5), "", alt, wfill)
                col_idx = 6
                if self.config.rpe_column:
                    self._apply_data(ws.cell(row=current_row, column=col_idx), "", alt, wfill)
                    col_idx += 1
                if self.config.notes_column:
                    self._apply_data(ws.cell(row=current_row, column=col_idx), "", alt, wfill)
                    col_idx += 1
                # Form cue
                cue_cell = ws.cell(row=current_row, column=col_idx)
                self._apply_data(cue_cell, ex.form_cue, alt, wfill)
                cue_cell.alignment = Alignment(horizontal="left", wrap_text=True)
                col_idx += 1
                if self.config.include_cardio:
                    self._apply_data(ws.cell(row=current_row, column=col_idx),
                                     "—" if not ex.is_cardio else "", alt, wfill)
                current_row += 1

            # Cardio day section
            if self.config.include_cardio:
                cardio_ex = [ex for ex in day.exercises if ex.is_cardio]
                if cardio_ex:
                    self._apply_day_header(
                        ws.cell(row=current_row, column=1), "— CARDIO —"
                    )
                    ws.merge_cells(f"A{current_row}:{merge_end}{current_row}")
                    current_row += 1
                    for j, h in enumerate(
                        ["Activity", "Duration", "Intensity", "Done", "Notes"], 1
                    ):
                        self._apply_header(ws.cell(row=current_row, column=j), h)
                    current_row += 1
                    for ex in cardio_ex:
                        ws.cell(row=current_row, column=1, value=ex.name).border = self._border
                        ws.cell(row=current_row, column=2, value=ex.weight_format).border = self._border
                        ws.cell(row=current_row, column=3, value=ex.form_cue).border = self._border
                        ws.cell(row=current_row, column=4, value="☐").border = self._border
                        ws.cell(row=current_row, column=5, value="").border = self._border
                        current_row += 1

            current_row += 1  # blank row between days

        # Wildcard slot
        if (
            self.config.include_wildcard_slots
            and self.config.wildcard_frequency > 0
            and (week % self.config.wildcard_frequency == 0)
        ):
            ws.cell(row=current_row, column=1,
                    value="— WILDCARD SLOT — (unplanned session)")
            ws.cell(row=current_row, column=1).font = Font(
                bold=True, italic=True, color="4040A0"
            )
            ws.cell(row=current_row, column=1).fill = _make_fill(self.theme.wildcard_bg)
            ws.merge_cells(f"A{current_row}:{merge_end}{current_row}")
            current_row += 1
            for j, h in enumerate(col_headers, 1):
                self._apply_header(ws.cell(row=current_row, column=j), h)
                ws.cell(row=current_row, column=j).fill = _make_fill(self.theme.wildcard_bg)
            current_row += 1
            for _ in range(3):
                for col in range(1, num_cols + 1):
                    ws.cell(row=current_row, column=col).border = self._border
                    ws.cell(row=current_row, column=col).fill = _make_fill(self.theme.wildcard_bg)
                current_row += 1
            current_row += 1

        # Weekly notes section
        ws.cell(row=current_row, column=1, value="WEEKLY NOTES:")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=11, name=t.body_font)
        current_row += 1
        for label in ["Bodyweight:", "Sleep (avg hrs):", "Energy (1–10):"]:
            ws.cell(row=current_row, column=1, value=label)
            ws.cell(row=current_row, column=2).border = self._border
            current_row += 1
        ws.cell(row=current_row, column=1, value="Notes:")
        ws.merge_cells(f"B{current_row}:{get_column_letter(min(num_cols, 6))}{current_row}")
        ws.cell(row=current_row, column=2).border = self._border

        # Column widths
        col_widths = [22, 12, 14, 6, 12]
        if self.config.rpe_column:
            col_widths.append(6)
        if self.config.notes_column:
            col_widths.append(32)
        col_widths.append(40)  # form cue
        if self.config.include_cardio:
            col_widths.append(12)
        for j, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = w

    def _build_rules_sheet(self):
        if not self.config.rules_text:
            return
        ws = self.wb.create_sheet("Rules")
        ws["A1"] = "PROGRAM RULES & METHODS"
        ws["A1"].font = Font(bold=True, size=14, name=self.theme.title_font)
        for i, line in enumerate(self.config.rules_text.strip().split("\n"), 3):
            cell = ws.cell(row=i, column=1, value=line)
            if line.strip().endswith(":") or line.strip().isupper():
                cell.font = Font(bold=True, name=self.theme.body_font)
        ws.column_dimensions["A"].width = 90

    def build(self, filename: Optional[str] = None) -> str:
        """Build the workbook and save to file. Returns the output path."""
        os.makedirs(self.config.output_path, exist_ok=True)
        if filename is None:
            safe = re.sub(r"[^\w\- ]", "", self.config.name).replace(" ", "_").lower()
            filename = f"{safe}.xlsx"
        if not os.path.isabs(filename):
            filename = os.path.join(self.config.output_path, filename)

        self._build_overview()
        self._build_progression_sheet()
        for week in range(1, self.config.total_weeks + 1):
            self._build_week_tab(week)
        self._build_rules_sheet()

        self.wb.save(filename)
        print(f"OK  Program saved: {filename}")
        return filename

# ─────────────────────────────────────────────────────────────────────────────
# PROGRAM MODIFIER
# ─────────────────────────────────────────────────────────────────────────────

class ProgramModifier:
    """Patch an existing Excel program file without recreating it. Methods chain."""

    def __init__(self, path: str):
        self.path = path
        self.wb = load_workbook(path)

    # ── Internal helpers ──────────────────────────────────────────────────────

    def _week_sheets(self, week_range: Optional[tuple[int, int]]):
        lo, hi = week_range or (1, 9999)
        for name in self.wb.sheetnames:
            n = _parse_week_num(name)
            if n is not None and lo <= n <= hi:
                yield n, self.wb[name]

    @staticmethod
    def _parse_weight(val: str) -> Optional[float]:
        if val is None:
            return None
        s = str(val).strip().lstrip("+").lower().replace("kg", "").strip()
        try:
            return float(s)
        except ValueError:
            return None

    @staticmethod
    def _is_exercise_row(row_cells) -> bool:
        first = str(row_cells[0].value or "").strip()
        if not first:
            return False
        low = first.lower()
        skip = ("day ", "session ", "día ", "exercise", "weekly notes",
                "bodyweight", "sleep", "energy", "notes:", "— wildcard",
                "— cardio", "week ")
        return not any(low.startswith(s) for s in skip)

    # ── Operations ────────────────────────────────────────────────────────────

    def scale_weights(
        self,
        week_range: Optional[tuple[int, int]] = None,
        pct: float = 90.0,
        rounding: float = 2.5,
    ) -> "ProgramModifier":
        """Scale all numeric weights in the given week range to pct%."""
        factor = pct / 100.0
        for _, ws in self._week_sheets(week_range):
            _, header_map = _find_header_row(ws)
            w_col = header_map.get("weight")
            if w_col is None:
                continue
            for row_cells in ws.iter_rows():
                if not self._is_exercise_row(row_cells):
                    continue
                cell = row_cells[w_col - 1]
                raw = str(cell.value or "")
                prefix = "+" if raw.strip().startswith("+") else ""
                kg = self._parse_weight(raw)
                if kg is not None and kg > 0:
                    new_val = round((kg * factor) / rounding) * rounding
                    cell.value = f"{prefix}{new_val:.4g}kg"
        return self

    def change_rep_scheme(
        self,
        week_range: Optional[tuple[int, int]] = None,
        exercise: str = "",
        new_scheme: str = "",
    ) -> "ProgramModifier":
        """Change the rep scheme for a specific exercise across a week range."""
        ex_lower = exercise.lower()
        for _, ws in self._week_sheets(week_range):
            _, header_map = _find_header_row(ws)
            sr_col = header_map.get("sets_reps")
            ex_col = header_map.get("exercise", 1)
            if sr_col is None:
                continue
            for row_cells in ws.iter_rows():
                if not self._is_exercise_row(row_cells):
                    continue
                name = str(row_cells[ex_col - 1].value or "").lower()
                if ex_lower in name:
                    row_cells[sr_col - 1].value = new_scheme
        return self

    def swap_exercise(
        self,
        old: str,
        new: str,
        week_range: Optional[tuple[int, int]] = None,
    ) -> "ProgramModifier":
        """Rename an exercise across all matching week tabs."""
        old_lower = old.lower()
        for _, ws in self._week_sheets(week_range):
            _, header_map = _find_header_row(ws)
            ex_col = header_map.get("exercise", 1)
            for row_cells in ws.iter_rows():
                cell = row_cells[ex_col - 1]
                if old_lower in str(cell.value or "").lower():
                    cell.value = new
        return self

    def add_rpe_column(
        self, week_range: Optional[tuple[int, int]] = None
    ) -> "ProgramModifier":
        """Insert an RPE column after Actual if not already present."""
        for _, ws in self._week_sheets(week_range):
            hrow, header_map = _find_header_row(ws)
            if hrow is None or "rpe" in header_map:
                continue
            actual_col = header_map.get("actual", max(header_map.values()))
            insert_at = actual_col + 1
            ws.insert_cols(insert_at)
            src = ws.cell(row=hrow, column=actual_col)
            rpe_cell = ws.cell(row=hrow, column=insert_at, value="RPE")
            # Copy header style from adjacent cell
            if src.font:
                rpe_cell.font = Font(
                    bold=src.font.bold,
                    color=(src.font.color.rgb
                           if src.font.color and src.font.color.type == "rgb" else "FFFFFF"),
                    size=src.font.size, name=src.font.name,
                )
            if src.fill and src.fill.fgColor and src.fill.fgColor.type == "rgb":
                rpe_cell.fill = PatternFill("solid", fgColor=src.fill.fgColor.rgb)
            rpe_cell.alignment = Alignment(horizontal="center")
            rpe_cell.border = src.border
        return self

    def add_date_column(
        self,
        week_range: Optional[tuple[int, int]] = None,
        start_date: Optional[date] = None,
        day_offsets: Optional[list[int]] = None,
    ) -> "ProgramModifier":
        """Append date text to each day header row within the week range."""
        if day_offsets is None:
            day_offsets = [0, 1, 3, 4]
        for week_num, ws in self._week_sheets(week_range):
            day_index = 0
            for row_cells in ws.iter_rows():
                first = str(row_cells[0].value or "").strip().lower()
                if first.startswith(("day ", "session ", "día ")):
                    cell = row_cells[0]
                    current_text = str(cell.value or "")
                    if "date:" in current_text.lower():
                        day_index += 1
                        continue
                    if start_date:
                        offset = day_offsets[day_index] if day_index < len(day_offsets) else day_index
                        d = start_date + timedelta(weeks=week_num - 1) + timedelta(days=offset)
                        date_str = d.isoformat()
                    else:
                        date_str = "________"
                    cell.value = current_text + f"  |  Date: {date_str}"
                    day_index += 1
        return self

    def add_wildcard_slot(
        self,
        after_week: int,
        num_rows: int = 3,
    ) -> "ProgramModifier":
        """Append a wildcard (surprise session) block at the bottom of a week tab."""
        name = f"Week {after_week}"
        if name not in self.wb.sheetnames:
            return self
        ws = self.wb[name]
        _, header_map = _find_header_row(ws)
        num_cols = max(header_map.values()) if header_map else 6

        # Find the last used row
        last_row = ws.max_row + 2
        wc_fill = _make_fill("E8E8FF")

        ws.cell(row=last_row, column=1,
                value="— WILDCARD SLOT — (unplanned session)").fill = wc_fill
        try:
            ws.merge_cells(f"A{last_row}:{get_column_letter(num_cols)}{last_row}")
        except Exception:
            pass
        last_row += 1

        # Header row
        for j, k in enumerate(header_map.items(), 1):
            c = ws.cell(row=last_row, column=k[1], value=k[0].replace("_", " ").title())
            c.fill = wc_fill
            c.border = _make_border("thin")
        last_row += 1

        for _ in range(num_rows):
            for col in range(1, num_cols + 1):
                c = ws.cell(row=last_row, column=col)
                c.border = _make_border("thin")
                c.fill = wc_fill
            last_row += 1
        return self

    def retheme(
        self,
        theme: str,
        week_range: Optional[tuple[int, int]] = None,
    ) -> "ProgramModifier":
        """Apply a new color theme to header rows in the given week range."""
        t = THEMES.get(theme)
        if not t:
            print(f"[Modifier] Unknown theme '{theme}'. Available: {list(THEMES)}")
            return self
        header_fill = _make_fill(t.header_bg)
        header_font_color = t.header_fg
        day_fill = _make_fill(t.day_bg)
        day_font_color = t.day_fg

        for _, ws in self._week_sheets(week_range):
            hrow, header_map = _find_header_row(ws)
            for row_cells in ws.iter_rows():
                first = str(row_cells[0].value or "").strip().lower()
                row_num = row_cells[0].row
                if first.startswith(("day ", "session ", "día ")):
                    for cell in row_cells:
                        if cell.value is not None:
                            cell.fill = day_fill
                            cell.font = Font(
                                bold=True, color=day_font_color,
                                size=t.body_size + 1, name=t.body_font
                            )
                elif hrow and row_num == hrow:
                    for cell in row_cells:
                        if cell.value is not None:
                            cell.fill = header_fill
                            cell.font = Font(
                                bold=True, color=header_font_color,
                                size=t.body_size, name=t.body_font
                            )
        return self

    def save(self, path: Optional[str] = None) -> "ProgramModifier":
        """Save the workbook. Defaults to overwriting the original file."""
        out = path or self.path
        self.wb.save(out)
        print(f"OK  Modified program saved: {out}")
        return self

# ─────────────────────────────────────────────────────────────────────────────
# BLOCK SCHEME TABLES  (Nacho's 30-week program)
# ─────────────────────────────────────────────────────────────────────────────

_B1 = {
    "squat": "3x5", "front_sq": "3x5", "bench": "3x5", "ohp": "3x5",
    "deadlift": "1x5", "clean": "7x2", "shrugs": "3x5",
    "chinup": "3x5", "pullup": "3x5", "dips": "2x6",
    "curl": "3x6", "row_str": "3x6", "row_pwr": "2x6", "tbar": "2x8",
    "rdl": "1x8", "lunges": "2x6/leg", "triceps": "2x10", "hammer": "2x10",
    "squat_vol": "3x6", "bench_bilbo": "1xAMRAP", "bench_vol": "3x8",
    "ohp_density": "4 min density",
}
_B2 = {
    "squat": "5x4", "front_sq": "5x4", "bench": "5x4", "ohp": "5x4",
    "deadlift": "1x5", "clean": "8x2", "shrugs": "4x4",
    "chinup": "4x4", "pullup": "3x5", "dips": "3x5",
    "curl": "3x8", "row_str": "3x8", "row_pwr": "3x6", "tbar": "3x8",
    "rdl": "1x8", "lunges": "2x8/leg", "triceps": "3x10", "hammer": "2x12",
    "squat_vol": "3x8", "bench_bilbo": "1xAMRAP", "bench_vol": "3x10",
    "ohp_density": "7 min density",
}
_B3 = {
    "squat": "5x3", "front_sq": "4x3", "bench": "5x3", "ohp": "5x3",
    "deadlift": "1x5", "clean": "8x2", "shrugs": "5x3",
    "chinup": "4x4", "pullup": "4x4", "dips": "3x5",
    "curl": "4x5", "row_str": "4x5", "row_pwr": "3x5", "tbar": "3x6",
    "rdl": "1x8", "lunges": "2x6/leg", "triceps": "3x8", "hammer": "3x10",
    "squat_vol": "2x8", "bench_bilbo": "1xAMRAP", "bench_vol": "3x8",
    "ohp_density": "30 reps / 9 min",
}
_B4 = {
    "squat": "3-2-1 wave ×2", "front_sq": "4x3", "bench": "3-2-1 wave ×2", "ohp": "3-2-1 wave ×2",
    "deadlift": "1x5", "clean": "8x2", "shrugs": "5x3",
    "chinup": "4x4", "pullup": "4x4", "dips": "3x6",
    "curl": "5 min density", "row_str": "3x6", "row_pwr": "3x5", "tbar": "3x6",
    "rdl": "1x8", "lunges": "3x5/leg", "triceps": "3x8", "hammer": "3x10",
    "squat_vol": "2x6", "bench_bilbo": "1xAMRAP", "bench_vol": "4x6",
    "ohp_density": "35 reps / 8 min",
}
_B5 = {
    "squat": "4x5", "front_sq": "4x5", "bench": "4x5", "ohp": "4x5",
    "deadlift": "1x5", "clean": "6x2", "shrugs": "4x5",
    "chinup": "3x6", "pullup": "3x6", "dips": "3x6",
    "curl": "3x8", "row_str": "3x6", "row_pwr": "3x5", "tbar": "3x6",
    "rdl": "1x8", "lunges": "2x6/leg", "triceps": "3x10", "hammer": "2x12",
    "squat_vol": "2x8", "bench_bilbo": "1xAMRAP", "bench_vol": "3x8",
    "ohp_density": "Giant: 40 reps",
}
_B6 = {
    "squat": "3x5", "front_sq": "3x5", "bench": "3x5", "ohp": "3x5",
    "deadlift": "1x5", "clean": "6x2", "shrugs": "3x5",
    "chinup": "3x5", "pullup": "3x5", "dips": "3x5",
    "curl": "3x5", "row_str": "3x5", "row_pwr": "3x5", "tbar": "3x6",
    "rdl": "1x8", "lunges": "2x6/leg", "triceps": "3x8", "hammer": "2x10",
    "squat_vol": "2x6", "bench_bilbo": "1xAMRAP", "bench_vol": "3x6",
    "ohp_density": "—",
}
_PULLBACK_SCHEMES = {
    "squat": "3x5", "front_sq": "3x5", "bench": "3x5", "ohp": "3x5",
    "deadlift": "1x5", "clean": "5x2", "shrugs": "3x5",
    "chinup": "3x5", "pullup": "3x5", "dips": "2x6",
    "curl": "3x6", "row_str": "3x6", "row_pwr": "2x6", "tbar": "2x8",
    "rdl": "1x8", "lunges": "2x6/leg", "triceps": "2x10", "hammer": "2x10",
    "squat_vol": "2x8", "bench_bilbo": "1xAMRAP", "bench_vol": "2x10",
    "ohp_density": "4 min",
}

_FORM_CUES = {
    "squat":    ["Brace hard. Fight chest drop. Own the weight.",
                 "More volume, same focus. Keep chest up.",
                 "Heavier now. Speed out of hole. Aggressive.",
                 "Wave loading. Each wave builds. Trust it.",
                 "Heavy 5s. Grind if needed. You're strong.",
                 "Peak 5RM. This is what you built. Execute."],
    "front_sq": ["Elbows HIGH. Pause if needed. Stay upright.",
                 "Volume building. Elbows never drop.",
                 "Heavier. If elbows drop, rack it.",
                 "Skill work. Perfect positions.",
                 "Heavy 5s. Fight for every rep.",
                 "Show your strength. Elbows up."],
    "bench":    ["Arch, retract, leg drive. Build the pattern.",
                 "More sets. Same tightness every rep.",
                 "Heavy. Pause on chest. Explode.",
                 "Waves. Stay tight between reps.",
                 "Heavy 5s. Control down, drive up.",
                 "105kg × 5. You've earned this."],
    "ohp":      ["Squeeze glutes. Bar path straight. Lockout.",
                 "Volume. Don't rush reps. Full lockout.",
                 "Heavy. Head through at top. Tight core.",
                 "Waves. Reset between reps if needed.",
                 "Heavy 5s. Grind the last reps.",
                 "65kg × 5. Stand tall. You got this."],
    "deadlift": ["Slack out first. Push floor away. Patience.",
                 "Same weight range. Perfect every rep.",
                 "Heavier. Bar drags legs. Lockout with glutes.",
                 "Heavy singles feel easy now.",
                 "Building to big pulls.",
                 "Show your strength. One set. All out."],
    "clean":    ["Jump, shrug, fast elbows. No arm pull.",
                 "More sets. Keep it explosive.",
                 "Heavier. Same speed or don't lift it.",
                 "Power and speed. Quality reps only.",
                 "Dialed in. Explosive every rep.",
                 "Show your power."],
    "shrugs":   ["Explosive. Traps to ears. Heavy.",
                 "More sets. Same explosion.",
                 "Heavier. Use straps. Maximum load.",
                 "Building that yoke.",
                 "Heavy and explosive. Yoke growing.",
                 "Massive traps. Show them."],
    "row_pwr":  ["Explosive pull. Some body english OK.",
                 "More power. Drive through.",
                 "Heavy and fast. Upper back thickness.",
                 "Explosive. Build that yoke.",
                 "Power reps. Feel strong.",
                 "Show the power."],
}


def _cue(lift_key: str, block_idx: int) -> str:
    cues = _FORM_CUES.get(lift_key, [])
    if cues and block_idx < len(cues):
        return cues[block_idx]
    return ""


def _ex(name, key, start, inc, fmt="kg", ref_lift=None, ref_pct=1.0,
        override_scheme=None, optional=False) -> ExerciseConfig:
    """Shorthand factory for ExerciseConfig with block-indexed form cues."""
    return ExerciseConfig(
        name=name, lift_key=key,
        start_weight=start, increment=inc,
        weight_format=fmt,
        ref_lift=ref_lift, ref_pct=ref_pct,
        override_scheme=override_scheme,
        is_optional=optional,
    )


# ─────────────────────────────────────────────────────────────────────────────
# DEFAULT CONFIG  (Nacho's 30-week program — form cues injected at build time)
# ─────────────────────────────────────────────────────────────────────────────

def _build_default_config() -> ProgramConfig:
    """Return the full 30-week program config."""
    # Form cues will be wired in _patch_form_cues after config is created
    days = [
        DayConfig(
            label="Day 1", focus="Squat + Bench Heavy", est_time="~50 min",
            exercises=[
                _ex("Squat",       "squat",    85,   1.5),
                _ex("Lunges",      "lunges",   30,   0.75),
                _ex("Bench Press", "bench",    75,   1.25),
                _ex("Dips",        "dips",     10,   1.0,  fmt="+kg"),
            ],
        ),
        DayConfig(
            label="Day 2", focus="Power + Posterior (Yoke)", est_time="~50 min",
            exercises=[
                _ex("Power Clean",   "clean",    52.5, 1.0),
                _ex("Deadlift",      "deadlift", 130,  2.5),
                _ex("Power Shrugs",  "shrugs",   90,   2.0),
                _ex("Pull-ups",      "pullup",   10,   1.0, fmt="+kg"),
                _ex("BB Row Power",  "row_pwr",  70,   1.0),
            ],
        ),
        DayConfig(
            label="Day 3", focus="Front Squat + OHP + Pull", est_time="~55–60 min",
            exercises=[
                _ex("Front Squat",  "front_sq", 65,   1.0),
                ExerciseConfig(name="Squat (Volume)", lift_key="squat_vol",
                               ref_lift="squat", ref_pct=0.75),
                _ex("OHP",          "ohp",      52.5, 0.5),
                _ex("BB Row Strict","row_str",  50,   1.0),
                _ex("T-Bar Row",    "tbar",     20,   1.0, fmt="+kg"),
                _ex("BB Curl",      "curl",     35,   0.6),
                _ex("RDL",          "rdl",      75,   0.5),
            ],
        ),
        DayConfig(
            label="Day 4", focus="Bench Volume + OHP Density + Arms", est_time="~55–60 min",
            exercises=[
                ExerciseConfig(name="Bench Bilbo",        lift_key="bench_bilbo",
                               ref_lift="bench", ref_pct=0.5,
                               override_scheme="1xAMRAP"),
                ExerciseConfig(name="Bench Volume",       lift_key="bench_vol",
                               ref_lift="bench", ref_pct=0.7),
                ExerciseConfig(name="OHP Density/Giant",  lift_key="ohp_density",
                               ref_lift="ohp",   ref_pct=0.85),
                _ex("Chin-ups",      "chinup",  10,   1.0, fmt="+kg"),
                _ex("BB Curl",       "curl",    35,   0.6),
                _ex("Triceps Ext",   "triceps", 15,   0.5, fmt="+kg"),
                _ex("Hammer Curl",   "hammer",  12.5, 0.4),
                ExerciseConfig(name="Nordic Curl (opt)",  lift_key="nordic",
                               override_scheme="2–3×5 eccentric",
                               weight_format="bw", is_optional=True,
                               form_cue="Slow eccentric. Control every rep."),
            ],
        ),
    ]

    blocks = [
        BlockConfig(start_week=1,  end_week=5,  label="Block 1 — Linear Base",        schemes=_B1),
        BlockConfig(start_week=6,  end_week=10, label="Block 2 — Volume Accumulation", schemes=_B2),
        BlockConfig(start_week=11, end_week=15, label="Block 3 — Intensity",           schemes=_B3),
        BlockConfig(start_week=16, end_week=20, label="Block 4 — Strength-Skill",      schemes=_B4),
        BlockConfig(start_week=21, end_week=25, label="Block 5 — 5RM Building",        schemes=_B5),
        BlockConfig(start_week=26, end_week=30, label="Block 6 — 5RM Peak & Test",     schemes=_B6),
    ]

    goals = [
        {"lift": "Squat",      "start": "85kg × 5",   "goal": "120kg × 5", "gain": "+35kg"},
        {"lift": "Bench Press","start": "75kg × 5",   "goal": "105kg × 5", "gain": "+30kg"},
        {"lift": "OHP",        "start": "52.5kg × 5", "goal": "65kg × 5",  "gain": "+12.5kg"},
        {"lift": "BB Curl",    "start": "35kg × 5",   "goal": "50kg × 5",  "gain": "+15kg"},
        {"lift": "Deadlift",   "start": "130kg × 5",  "goal": "LP",        "gain": "—"},
        {"lift": "Front Squat","start": "65kg × 5",   "goal": "LP",        "gain": "—"},
        {"lift": "Chin-ups",   "start": "+10kg × 5",  "goal": "LP (aggressive)", "gain": "—"},
        {"lift": "Pull-ups",   "start": "+10kg × 5",  "goal": "LP (aggressive)", "gain": "—"},
        {"lift": "Power Shrugs","start": "90kg",       "goal": "LP (aggressive)", "gain": "—"},
    ]

    philosophy = (
        "High frequency, low volume, high intensity. Build 5RM strength, not peaking for 1RM.\n"
        "Pullback every 5th week. 6 blocks of progression. Week 30: TEST 5RM."
    )

    rules_text = """
PROGRESSION RULES:

1. ADD WEIGHT WHEN YOU HIT ALL REPS
   Hit all reps with good form → add weight next week.
   Missed reps → repeat same weight.
   Missed twice → drop 10% and rebuild.

2. PULLBACK WEEKS (5, 10, 15, 20, 25)
   Reduce all weights ~10%. Same sets and reps. Recovery + consolidation.
   Come back stronger next week.

3. NEVER SKIP MAIN LIFTS
   Squat, Front Squat, Bench, OHP, Deadlift = non-negotiable.
   Short on time? Cut accessories, keep main lifts.

4. LISTEN TO YOUR BODY
   Joint pain = reduce weight, check form.
   Exhausted = hit main lifts only.
   Feeling strong = still follow the program (don't jump ahead).

METHODS EXPLAINED:

LINEAR PROGRESSION (Blocks 1, 5, 6):
Add weight every week. Simple and effective. 3x5 or 4x5.

VOLUME ACCUMULATION (Block 2):
Same weight, more sets. 4x4 → 5x4. Builds muscle and work capacity.

INTENSITY BLOCK (Block 3):
Heavier weights, fewer reps. 5x3 → 6x2. Builds strength.

WAVE LOADING (Block 4):
3-2-1 waves: 3 reps, rest, 2 reps, rest, 1 rep = 1 wave. Do 2 waves.
Add weight each wave if possible.

BILBO METHOD (Bench Day 4):
50% of estimated 1RM. ONE set to absolute failure. Add 2.5kg every week.

DENSITY SETS (OHP Day 4, Curls Block 4):
Fixed weight, fixed time. Max reps. Rest as needed within window.
Progression: add 30 seconds or add reps.

GIANT SETS (Block 5 OHP):
Fixed total reps (e.g. 40). Complete as fast as possible.
Progression: same reps, less time — or more reps.

REST PERIODS:
Main lifts: 3–4 min. Power Clean: 60–90 sec (keep explosive). Accessories: 90 sec–2 min.

WEEK 30 TEST DAY PROTOCOL:
Day 1: Squat 120kg × 5, Bench 105kg × 5.
Day 2: Deadlift PR attempt.
Day 3: Front Squat PR, OHP 65kg × 5.
Day 4: BB Curl 50kg × 5, remaining PRs.
Warm up thoroughly. Film your lifts. Celebrate.
""".strip()

    return ProgramConfig(
        name="30-Week Strength Program",
        total_weeks=30,
        days=days,
        blocks=blocks,
        pullback_weeks=[5, 10, 15, 20, 25],
        pullback_schemes=_PULLBACK_SCHEMES,
        progression_method="BLOCK_PERIODIZATION",
        theme="professional_blue",
        rpe_column=True,
        notes_column=True,
        include_wildcard_slots=False,
        pullback_pct=0.90,
        rounding=2.5,
        goals=goals,
        philosophy=philosophy,
        rules_text=rules_text,
    )


def _inject_form_cues(config: ProgramConfig):
    """Inject block-indexed form cues into exercise objects (requires blocks to be built)."""
    for day in config.days:
        for ex in day.exercises:
            if not ex.form_cue and ex.lift_key in _FORM_CUES:
                # Use block 0 cue as default (will show correctly for week 1)
                ex.form_cue = _FORM_CUES[ex.lift_key][0]


# ─────────────────────────────────────────────────────────────────────────────
# DEMO CONFIG  (12-week hypertrophy, clean_forest theme)
# ─────────────────────────────────────────────────────────────────────────────

def _build_demo_config(theme: str = "professional_blue") -> ProgramConfig:
    """12-week hypertrophy demo program for theme previewing."""
    return ProgramConfig(
        name="12-Week Hypertrophy Block",
        total_weeks=12,
        progression_method="DOUBLE_PROGRESSION",
        theme=theme,
        rpe_column=True,
        notes_column=True,
        include_wildcard_slots=True,
        wildcard_frequency=4,
        pullback_weeks=[4, 8, 12],
        pullback_pct=0.85,
        rounding=2.5,
        philosophy="Hypertrophy focus. Rep ranges 8–12. Double progression.",
        goals=[
            {"lift": "Bench Press", "start": "60kg × 8", "goal": "80kg × 8", "gain": "+20kg"},
            {"lift": "Squat",       "start": "70kg × 8", "goal": "95kg × 8", "gain": "+25kg"},
            {"lift": "BB Row",      "start": "55kg × 8", "goal": "70kg × 8", "gain": "+15kg"},
        ],
        blocks=[
            BlockConfig(1, 4,  "Accumulation",  {"squat": "3x8-12", "bench": "3x8-12", "row": "3x8-12"}),
            BlockConfig(5, 8,  "Intensification",{"squat": "4x6-10", "bench": "4x6-10", "row": "4x6-10"}),
            BlockConfig(9, 12, "Realization",    {"squat": "4x5-8",  "bench": "4x5-8",  "row": "4x5-8"}),
        ],
        days=[
            DayConfig("Day 1", "Upper Push", "~45 min", [
                ExerciseConfig("Bench Press",     "bench",  60, 2.0, override_scheme="3x8-12"),
                ExerciseConfig("Incline DB Press","incline",22, 1.0, override_scheme="3x10-15"),
                ExerciseConfig("OHP",             "ohp",    40, 1.0, override_scheme="3x10-12"),
                ExerciseConfig("Lateral Raises",  "lateral", 8, 0.5, override_scheme="3x15-20",
                               weight_format="kg"),
            ]),
            DayConfig("Day 2", "Lower Pull", "~45 min", [
                ExerciseConfig("Squat",      "squat",    70, 2.5, override_scheme="3x8-12"),
                ExerciseConfig("RDL",        "rdl",      60, 1.5, override_scheme="3x10-12"),
                ExerciseConfig("Leg Curl",   "leg_curl", 40, 1.0, override_scheme="3x12-15"),
                ExerciseConfig("Calf Raise", "calf",     50, 1.0, override_scheme="4x15-20"),
            ]),
            DayConfig("Day 3", "Upper Pull + Arms", "~50 min", [
                ExerciseConfig("BB Row",      "row",    55, 1.5, override_scheme="3x8-12"),
                ExerciseConfig("Lat Pulldown","pulldown",50, 1.0, override_scheme="3x10-12"),
                ExerciseConfig("BB Curl",     "curl",   30, 0.75, override_scheme="3x10-15"),
                ExerciseConfig("Triceps Ext", "triceps",20, 0.5,  override_scheme="3x12-15"),
            ]),
        ],
    )

# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Program Builder V2 — flexible Excel strength program generator"
    )
    parser.add_argument(
        "--demo", action="store_true",
        help="Generate a 12-week hypertrophy sample in all 4 themes",
    )
    parser.add_argument(
        "--extend-current", metavar="FILE",
        help="Add RPE + Date columns to an existing .xlsx file (produces FILE_v2.xlsx)",
    )
    parser.add_argument(
        "--theme", default="professional_blue",
        choices=list(THEMES),
        help="Theme for the default 30-week program (default: professional_blue)",
    )
    parser.add_argument(
        "--output", default=".",
        help="Output directory (default: current directory)",
    )
    parser.add_argument(
        "--start-date", metavar="YYYY-MM-DD",
        help="Program start date — auto-fills session dates",
    )
    args = parser.parse_args()

    if args.demo:
        print("Generating demo programs in all 4 themes...")
        os.makedirs(args.output, exist_ok=True)
        for theme_name in THEMES:
            cfg = _build_demo_config(theme=theme_name)
            cfg.output_path = args.output
            out = os.path.join(args.output, f"demo_12week_{theme_name}.xlsx")
            ProgramBuilder(cfg).build(out)
        print(f"\nOK  4 demo files written to: {os.path.abspath(args.output)}/")
        return

    if args.extend_current:
        path = args.extend_current
        if not os.path.isfile(path):
            print(f"Error: file not found: {path}")
            return
        base, ext = os.path.splitext(path)
        out = f"{base}_v2{ext}"
        start_d: Optional[date] = None
        if args.start_date:
            try:
                start_d = date.fromisoformat(args.start_date)
            except ValueError:
                print(f"Warning: invalid start-date '{args.start_date}', skipping date fill")
        (
            ProgramModifier(path)
            .add_rpe_column()
            .add_date_column(start_date=start_d)
            .save(out)
        )
        return

    # Default: regenerate Nacho's 30-week program
    cfg = _build_default_config()
    cfg.theme = args.theme
    cfg.output_path = args.output
    if args.start_date:
        try:
            cfg.start_date = date.fromisoformat(args.start_date)
        except ValueError:
            print(f"Warning: invalid start-date '{args.start_date}', skipping")
    _inject_form_cues(cfg)
    fname = os.path.join(args.output, "strength_30weeks.xlsx")
    ProgramBuilder(cfg).build(fname)


if __name__ == "__main__":
    main()
